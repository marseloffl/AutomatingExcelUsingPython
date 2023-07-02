from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('marks.xlsx')
ws = wb.active

# To create new sheet in workbook
#wb.create_sheet('Test')

# To show how many sheets in workbook
#print(wb.sheetnames)

# To print particular values from row & col
#for row in ws.iter_rows(min_row=1, max_row=5, min_col=3, max_col=6, values_only=True):
#    print(row)

# To Print Values from workbook
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)


# To merge & unmerge cells
#ws.merge_cells("A1:D1")
#ws.unmerge_cells("A1:D1")

# To add & delete row / column
#ws.insert_rows(7) - It will insert a row at row 7
#ws.delete_rows(7) - It will delete the rows at row 7

#ws.insert_cols(2) - It Adds new column at 2
#ws.delete_cols(2)

# To move the range of values
#ws.move_range("C1:D11", rows=2, cols=2) - if negative value in rows it moves up.
