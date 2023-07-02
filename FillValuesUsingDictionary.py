from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

marks = {
	"Mari": {
		"English": 65,
		"Maths": 78,
		"Science": 98,
		"Social": 89
	},
	"Raja": {
		"English": 55,
		"Maths": 72,
		"Science": 87,
		"Social": 95
	},
	"Prem": {
		"English": 100,
		"Maths": 45,
		"Science": 75,
		"Social": 92
	},
	"Logesh": {
		"English": 30,
		"Maths": 25,
		"Science": 45,
		"Social": 100
	},
	"Nallarusu": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(marks['Mari'].keys())
ws.append(headings)

for person in marks:
	grades = list(marks[person].values())
	ws.append([person] + grades)

# To Find Average Marks in each Subject
for col in range(2, len(marks['Mari']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(marks)}"

# To Style the heading Bold
for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True)

wb.save('NewResult.xlsx')